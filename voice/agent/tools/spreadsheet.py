"""
agent/tools/spreadsheet.py
Spreadsheet tools: read and update Excel/CSV files at the cell level
without overwriting the whole file.
"""
import csv
import logging
import os
import tempfile
from pathlib import Path

log = logging.getLogger("intellivox.spreadsheet")

HOME = str(Path.home())

SAFE_BASE_DIRS = [
    os.path.join(HOME, "Desktop"),
    os.path.join(HOME, "Documents"),
    os.path.join(HOME, "Downloads"),
    HOME,
]


def _is_safe_path(path: str) -> bool:
    abs_path = str(Path(path).expanduser().resolve())
    return any(abs_path.startswith(base) for base in SAFE_BASE_DIRS)


def _resolve(path: str) -> str:
    return str(Path(path).expanduser().resolve())


def _xls_to_xlsx(xls_path: str) -> str:
    """Convert a legacy .xls file to a temporary .xlsx. Caller must delete the temp file."""
    try:
        import xlrd
        import openpyxl
    except ImportError as e:
        raise RuntimeError(f"pip install xlrd openpyxl — {e}")

    wb_old = xlrd.open_workbook(xls_path)
    wb_new = openpyxl.Workbook()
    wb_new.remove(wb_new.active)

    for name in wb_old.sheet_names():
        ws_old = wb_old.sheet_by_name(name)
        ws_new = wb_new.create_sheet(title=name)
        for ri in range(ws_old.nrows):
            for ci in range(ws_old.ncols):
                ws_new.cell(row=ri + 1, column=ci + 1).value = ws_old.cell(ri, ci).value

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb_new.save(tmp.name)
    return tmp.name


def _needs_conversion(path: str) -> bool:
    """True if path is a legacy .xls (not .xlsx/.xlsm)."""
    name = path.lower()
    return name.endswith(".xls") and not name.endswith(".xlsx") and not name.endswith(".xlsm")


# ── Read ───────────────────────────────────────────────────────────────────────

def read_spreadsheet(path: str, sheet: str = None) -> dict:
    """
    Read an Excel (.xlsx/.xls) or CSV file.
    Returns: { success, headers, rows (list of dicts), row_count, path, sheet }
    rows[i] also includes a "__row_number__" key (1-based, header = row 1).
    """
    expanded = _resolve(path)
    if not os.path.exists(expanded):
        return {"success": False, "message": f"File not found: {path}"}
    if not _is_safe_path(expanded):
        return {"success": False, "message": "Access denied: outside allowed directories"}

    ext = Path(expanded).suffix.lower()

    if ext == ".csv":
        return _read_csv(expanded)
    elif ext in (".xlsx", ".xls", ".xlsm"):
        return _read_excel(expanded, sheet)
    else:
        return {"success": False, "message": f"Unsupported format: {ext}. Use .xlsx or .csv"}


def _read_csv(path: str) -> dict:
    try:
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            rows = []
            for i, row in enumerate(reader, start=2):
                r = dict(row)
                r["__row_number__"] = i
                rows.append(r)
        return {"success": True, "headers": list(headers), "rows": rows,
                "row_count": len(rows), "path": path, "sheet": None}
    except Exception as e:
        return {"success": False, "message": f"CSV read error: {e}"}


def _read_excel(path: str, sheet: str = None) -> dict:
    tmp_path = None
    try:
        if _needs_conversion(path):
            tmp_path = _xls_to_xlsx(path)
            xlsx_path = tmp_path
        else:
            xlsx_path = path

        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        rows_raw = list(ws.iter_rows(values_only=True))
        if not rows_raw:
            return {"success": True, "headers": [], "rows": [], "row_count": 0,
                    "path": path, "sheet": ws.title}
        headers = [str(h) if h is not None else f"Col{i+1}"
                   for i, h in enumerate(rows_raw[0])]
        rows = []
        for row_num, row_vals in enumerate(rows_raw[1:], start=2):
            r = {headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row_vals)}
            r["__row_number__"] = row_num
            rows.append(r)
        wb.close()
        return {"success": True, "headers": headers, "rows": rows,
                "row_count": len(rows), "path": path, "sheet": ws.title}
    except ImportError:
        return {"success": False, "message": "openpyxl not installed. Run: pip install openpyxl"}
    except RuntimeError as e:
        return {"success": False, "message": str(e)}
    except Exception as e:
        return {"success": False, "message": f"Excel read error: {e}"}
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass


# ── Update ─────────────────────────────────────────────────────────────────────

def update_spreadsheet(path: str, updates: list, sheet: str = None) -> dict:
    """
    Update specific cells in an Excel or CSV file without touching the rest.

    updates: list of dicts, each with ONE of these shapes:
      1. Direct cell update:
         { "row": 3, "column": "Invoice Amount", "value": "1500" }
         row is 1-based (row 1 = header row, row 2 = first data row).

      2. Find-then-update (search a column for a value, update another column):
         { "find_column": "Invoice No", "find_value": "INV-001",
           "set_column": "Amount", "set_value": "1500" }
         Leave find_value empty ("") to match ALL rows.

    Column names are matched case-insensitively.
    Returns: { success, updated_cells, path }
    """
    expanded = _resolve(path)
    if not os.path.exists(expanded):
        return {"success": False, "message": f"File not found: {path}"}
    if not _is_safe_path(expanded):
        return {"success": False, "message": "Access denied: outside allowed directories"}

    ext = Path(expanded).suffix.lower()

    if ext == ".csv":
        return _update_csv(expanded, updates)
    elif ext in (".xlsx", ".xls", ".xlsm"):
        return _update_excel(expanded, updates, sheet)
    else:
        return {"success": False, "message": f"Unsupported format: {ext}"}


def _update_csv(path: str, updates: list) -> dict:
    try:
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            rows = [dict(row) for row in reader]

        # Case-insensitive header lookup
        headers_lower = {h.lower(): h for h in headers}

        updated_cells = []

        for upd in updates:
            if "find_column" in upd:
                fc_key = upd["find_column"].lower()
                fc = headers_lower.get(fc_key, upd["find_column"])
                fv = str(upd.get("find_value", ""))
                sc_key = upd["set_column"].lower()
                sc = headers_lower.get(sc_key, upd["set_column"])
                sv = str(upd["set_value"])
                match_all = (fv == "")
                for i, row in enumerate(rows):
                    cell_val = str(row.get(fc, "")).strip()
                    if match_all or cell_val == fv.strip():
                        row[sc] = sv
                        updated_cells.append({"row": i + 2, "column": sc, "value": sv})
            elif "row" in upd:
                row_num = int(upd["row"]) - 2
                col_key = upd["column"].lower()
                col = headers_lower.get(col_key, upd["column"])
                val = str(upd["value"])
                if 0 <= row_num < len(rows):
                    rows[row_num][col] = val
                    updated_cells.append({"row": int(upd["row"]), "column": col, "value": val})

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)

        return {"success": True, "updated_cells": updated_cells,
                "path": path, "message": f"Updated {len(updated_cells)} cell(s)"}
    except Exception as e:
        return {"success": False, "message": f"CSV update error: {e}"}


def _update_excel(path: str, updates: list, sheet: str = None) -> dict:
    tmp_path = None
    try:
        if _needs_conversion(path):
            tmp_path = _xls_to_xlsx(path)
            # Auto-upgrade: save result as .xlsx alongside the original
            xlsx_path = str(Path(path).with_suffix(".xlsx"))
            import shutil
            shutil.copy2(tmp_path, xlsx_path)
            work_path = xlsx_path
            log.info("Auto-converted %s → %s", path, xlsx_path)
        else:
            work_path = path

        import openpyxl
        wb = openpyxl.load_workbook(work_path)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

        # Build case-insensitive header → column-index map from row 1
        headers = {str(cell.value).strip().lower(): cell.column
                   for cell in ws[1] if cell.value is not None}

        updated_cells = []

        for upd in updates:
            if "find_column" in upd:
                fc  = upd["find_column"].lower()
                fv  = str(upd.get("find_value", "")).strip()
                sc  = upd["set_column"].lower()
                sv  = str(upd["set_value"])
                fc_idx = headers.get(fc)
                sc_idx = headers.get(sc)
                if fc_idx is None:
                    log.warning("Column not found: %s", upd["find_column"])
                    continue
                if sc_idx is None:
                    log.warning("Column not found: %s", upd["set_column"])
                    continue
                match_all = (fv == "")
                for row in ws.iter_rows(min_row=2):
                    cell_val = row[fc_idx - 1].value
                    cell_str = str(cell_val).strip() if cell_val is not None else ""
                    if match_all or cell_str == fv:
                        row[sc_idx - 1].value = sv
                        updated_cells.append(
                            {"row": row[0].row, "column": upd["set_column"], "value": sv})
            elif "row" in upd:
                row_num  = int(upd["row"])
                col_name = upd["column"].lower()
                val      = upd["value"]
                col_idx  = headers.get(col_name)
                if col_idx is None:
                    log.warning("Column not found: %s", upd["column"])
                    continue
                ws.cell(row=row_num, column=col_idx).value = val
                updated_cells.append({"row": row_num, "column": upd["column"], "value": val})

        wb.save(work_path)
        wb.close()
        return {"success": True, "updated_cells": updated_cells,
                "path": work_path, "message": f"Updated {len(updated_cells)} cell(s)"}
    except ImportError:
        return {"success": False, "message": "openpyxl not installed. Run: pip install openpyxl"}
    except RuntimeError as e:
        return {"success": False, "message": str(e)}
    except Exception as e:
        return {"success": False, "message": f"Excel update error: {e}"}
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
